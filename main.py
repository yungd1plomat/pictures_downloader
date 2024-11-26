import requests
import os
import re
from urllib.parse import unquote, urlparse, parse_qs
from openpyxl import load_workbook
from unidecode import unidecode
from loguru import logger
from zipfile import ZipFile
from io import BytesIO

DATA_FOLDER = 'data/'
SAVE_FOLDER = 'result/'
IMAGES_FOLDER = os.path.join(SAVE_FOLDER, 'images')
EXTENSIONS = ['.png', '.jpg', '.jpeg', '.pdf', '.bmp']

excel_files = os.listdir(DATA_FOLDER)

if not os.path.exists(SAVE_FOLDER):
    os.mkdir(SAVE_FOLDER)

if not os.path.exists(IMAGES_FOLDER):
    os.mkdir(IMAGES_FOLDER)

def normalize_filename(filename):
    normalized = unidecode(filename)
    normalized = re.sub(r'[^\w.-]', '_', normalized)
    normalized = re.sub(r'_+', '_', normalized).strip('_')
    return normalized

def download_yandex_file(url, path = None):
    response = requests.get(f'https://cloud-api.yandex.net/v1/disk/public/resources/download?public_key={url}&path={path}')
    resp = response.json()
    if "error" in resp:
        logger.warning("Can't download file from yandex: " + resp["message"] + " " + url)
        return None
    download_link = resp['href']
    parsed_url = urlparse(download_link)
    query_params = parse_qs(parsed_url.query)
    filename = unquote(query_params.get('filename', [None])[0])
    filename = normalize_filename(filename)
    response = requests.get(download_link)
    with open(os.path.join(IMAGES_FOLDER, filename), 'wb') as f:
        f.write(response.content)
    return filename

def download_yandex_folder(url):
    response = requests.get(f'https://cloud-api.yandex.net/v1/disk/public/resources/download?public_key={url}')
    resp = response.json()
    if "error" in resp:
        logger.warning("Can't download file from yandex folder: " + resp["message"] + " " + url)
        return None
    download_link = resp['href']
    response = requests.get(download_link)
    with ZipFile(BytesIO(response.content)) as zip_ref:
        file_list = zip_ref.namelist()
        images = []
        for file_path in file_list:
            if any(file_path.lower().endswith(ext) for ext in EXTENSIONS):
                original_filename = os.path.basename(file_path)
                
                normalized_filename = normalize_filename(original_filename)

                with zip_ref.open(file_path) as source:
                    content = source.read()
                    output_path = os.path.join(IMAGES_FOLDER, normalized_filename)
                    
                    with open(output_path, 'wb') as target:
                        target.write(content)
                images.append(normalized_filename)
        return images

def download_raw_file(url):
    response = requests.get(url, allow_redirects=True)
    filename = url.split('/')[-1]
    filename = normalize_filename(filename)
    with open(os.path.join(IMAGES_FOLDER, filename), 'wb') as f:
        f.write(response.content)
    return filename

def process_cell(cell):
    raw_url = cell.hyperlink.target if cell.hyperlink else cell.value
    all_urls = re.split(r'[ \t\r\n,]+', raw_url)
    all_urls = [part for part in all_urls if part]
    for url in all_urls:
        url = unquote(url)
        if any(ext in url.lower() for ext in EXTENSIONS):
            if url.lower().startswith('https://disk.yandex.ru/d/'):
                folder = url.split('https://disk.yandex.ru/d/')[1].split('/')[0]
                path = url.split(folder)[1]
                folder_url = 'https://disk.yandex.ru/d/' + folder
                filename = download_yandex_file(folder_url, path)
                if filename:
                    cell.value = filename
                    cell.hyperlink = None
                    logger.info(f"Downloaded yandex file from folder: {filename}")
            else:
                filename = download_raw_file(url)
                if filename:
                    cell.value = filename
                    cell.hyperlink = None
                    logger.info(f"Downloaded file: {filename}")
        elif url.startswith('https://disk.yandex.ru/i/'):
            filename = download_yandex_file(url)
            if filename:
                cell.value = filename
                cell.hyperlink = None
                logger.info(f"Downloaded file from yandex disk: {filename}")
        elif url.startswith('https://disk.yandex.ru/d/'):
            filenames = download_yandex_folder(url)
            if filenames and len(filenames) > 0:
                cell.value = '\n'.join(filenames)
                cell.hyperlink = None
                logger.info(f"Downloaded folder from yandex disk: {', '.join(filenames)}")
        else:
            logger.warning(f'Unknown URL type: {url}')

def process_file(file_name):
    file_path = os.path.join(DATA_FOLDER, file_name)

    workbook = load_workbook(file_path)
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        cells_to_process = []
        for row in worksheet.iter_rows():
            for cell in row:
                if (cell.hyperlink is not None and cell.hyperlink.target.startswith("http")) or (cell.value is not None and str(cell.value).startswith("http")):
                    cells_to_process.append(cell)

        for cell in cells_to_process:
            process_cell(cell)
            workbook.save(os.path.join(SAVE_FOLDER, file_name))

if __name__ == "__main__":
    for file in excel_files:
        process_file(file)
